import os
import sys
import glob
import logging
import threading
from http.server import HTTPServer, BaseHTTPRequestHandler
from datetime import datetime
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

# Lightweight HTTP Health Check server for Render Free Web Service
class HealthCheckHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.send_header("Content-type", "text/plain")
        self.end_headers()
        self.wfile.write(b"Atsede Teguhan Bot is Running 24/7!")

    def log_message(self, format, *args):
        return  # Suppress health check logging

def start_health_server():
    port = int(os.environ.get("PORT", 8080))
    server = HTTPServer(("0.0.0.0", port), HealthCheckHandler)
    server.serve_forever()
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ConversationHandler,
    ContextTypes,
    filters,
)

# Setup logging
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)

BOT_TOKEN = "8706655732:AAHs877SHaGX7HnHVa80VAyafiA2GnqBPW8"
ADMIN_IDS = [1537845176]  # @Shetesfa (Your Telegram ID)
ADMIN_USERNAMES = ["shetesfa", "dawitmesenko", "abamekari"]  # Authorized Admins
WORKING_DIR = os.path.dirname(os.path.abspath(__file__))
SURVEY_EXCEL = os.path.join(WORKING_DIR, "survey_results.xlsx")

async def check_is_admin(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    """Checks if a user is an authorized admin by ID, username, or group admin status."""
    user = update.effective_user
    if not user:
        return False
    # 1. Match by Telegram ID
    if user.id in ADMIN_IDS:
        return True
    # 2. Match by Username
    if user.username and user.username.lower().replace("@", "") in ADMIN_USERNAMES:
        return True
    # 3. Match by Group Admin status in the group
    if update.effective_chat and update.effective_chat.type in ["group", "supergroup"]:
        try:
            member = await context.bot.get_chat_member(update.effective_chat.id, user.id)
            if member.status in ["creator", "administrator"]:
                return True
        except Exception:
            pass
    return False

# -------------------------------------------------------------
# 1. EXCEL DATABASE HELPER FOR SURVEY
# -------------------------------------------------------------
def init_survey_excel():
    if not os.path.exists(SURVEY_EXCEL):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Responses"
        ws.append(["ቀን እና ሰዓት (Timestamp)", "የተጠቃሚ መለያ (User ID)", "ሙሉ ስም (Full Name)", "የቴሌግራም ስም (Username)", "የተመረጠው ምርጫ (Choice)"])
        wb.save(SURVEY_EXCEL)

def record_vote(user_id, full_name, username, choice_text):
    """Saves vote only if user hasn't voted before. Returns (is_new, choice_text)."""
    init_survey_excel()
    wb = openpyxl.load_workbook(SURVEY_EXCEL)
    ws = wb.active

    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=2).value) == str(user_id):
            # User already voted!
            existing_choice = ws.cell(row=r, column=5).value
            return False, existing_choice

    # First time voting: save to Excel
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([now, user_id, full_name, username, choice_text])
    wb.save(SURVEY_EXCEL)
    return True, choice_text

# -------------------------------------------------------------
# 2. GROUP SURVEY HANDLERS
# -------------------------------------------------------------
async def cmd_ask(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sends the survey question with inline buttons to the group (Admins only)."""
    if not await check_is_admin(update, context):
        return  # Silently ignore non-admins

    survey_text = (
        "📢 **የአስተያየት መጠይቅ (Survey)**\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "የእለተ ዕሁድ ትምህርት 09:00 ሰዓት የነበረው በነበረው **09:00 - 11:00** ሰዓት ይቀጥል ወይስ ከ**10:00** ሰዓት ይጀምር እና **12:00** ሰዓት ይለቅ?\n\n"
        "1. በነበረው ይቀጥል ከ 09:00 - 11:00\n"
        "2. ወይስ ከ 10:00 - 12:00\n\n"
        "⚠️ _ማሳሰቢያ፦ ይህ ነገር መረጃ ለመሰብሰብ ነው እንጂ አይተገበርም።_\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "እባክዎ ከታች ካሉት አማራጮች አንዱን ይምረጡ፦"
    )

    keyboard = [
        [InlineKeyboardButton("1️⃣ በነበረው ይቀጥል (09:00 - 11:00)", callback_data="opt_1")],
        [InlineKeyboardButton("2️⃣ ከ 10:00 - 12:00 ይሁን", callback_data="opt_2")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(
        survey_text, reply_markup=reply_markup, parse_mode="Markdown"
    )

async def handle_survey_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handles button clicks on the survey with strict single-vote enforcement."""
    query = update.callback_query
    user = query.from_user

    choices = {
        "opt_1": "1. በነበረው ይቀጥል (09:00 - 11:00)",
        "opt_2": "2. ከ 10:00 - 12:00 ይሁን",
    }

    selected_choice = choices.get(query.data, "ያልታወቀ")
    username_str = f"@{user.username}" if user.username else "የለም (None)"

    # Check and record vote
    is_new, recorded_choice = record_vote(
        user_id=user.id,
        full_name=user.full_name or "Unknown",
        username=username_str,
        choice_text=selected_choice,
    )

    if not is_new:
        # User already voted! Block them and show warning popup
        await query.answer(
            text=f"⚠️ ይቅርታ! ከዚህ በፊት ድምጽ ሰጥተዋል።\nየመረጡት፦ {recorded_choice}\n(አንድ ሰው መምረጥ የሚችለው አንድ ጊዜ ብቻ ነው!)",
            show_alert=True,
        )
        return

    # First-time vote success
    await query.answer(
        text=f"✅ ምርጫዎ ተመዝግቧል!\nየመረጡት፦ {selected_choice}\nእናመሰግናለን!",
        show_alert=True,
    )

    # Send instant notification to Admin in private chat
    try:
        admin_alert = (
            f"🔔 **አዲስ ድምጽ ተመዝግቧል! (New Vote)**\n"
            f"👤 ስም፦ {user.full_name}\n"
            f"📱 Username፦ {username_str}\n"
            f"🗳️ የመረጡት፦ **{selected_choice}**"
        )
        for admin_id in ADMIN_IDS:
            await context.bot.send_message(chat_id=admin_id, text=admin_alert, parse_mode="Markdown")
    except Exception as e:
        logger.warning(f"Could not notify admin: {e}")

async def cmd_reset_survey(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Allows authorized admins to wipe all survey data and start fresh."""
    if not await check_is_admin(update, context):
        return  # Silently ignore non-admins

    init_survey_excel()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Responses"
    ws.append(["ቀን እና ሰዓት (Timestamp)", "የተጠቃሚ መለያ (User ID)", "ሙሉ ስም (Full Name)", "የቴሌግራም ስም (Username)", "የተመረጠው ምርጫ (Choice)"])
    wb.save(SURVEY_EXCEL)
    await update.message.reply_text("🧹 ሁሉም የድምጽ መረጃዎች ተሰርዘዋል! በአዲስ መጀመር ይችላሉ። (All survey data wiped clean!)")

async def cmd_survey_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Shows summary of votes and sends the Excel file to admin (Admins only)."""
    if not await check_is_admin(update, context):
        return  # Silently ignore non-admins

    init_survey_excel()
    wb = openpyxl.load_workbook(SURVEY_EXCEL)
    ws = wb.active

    total_votes = ws.max_row - 1
    if total_votes < 1:
        await update.message.reply_text("📊 እስካሁን ምንም የተመዘገበ ድምጽ የለም።")
        return

    count_opt1 = 0
    count_opt2 = 0
    for r in range(2, ws.max_row + 1):
        choice = str(ws.cell(row=r, column=5).value)
        if "09:00" in choice:
            count_opt1 += 1
        elif "10:00" in choice:
            count_opt2 += 1

    stats_text = (
        f"📊 **የአስተያየት ውጤት ማጠቃለያ**\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👥 አጠቃላይ የመረጡ ሰዎች፦ **{total_votes}**\n\n"
        f"1️⃣ በነበረው ይቀጥል (09:00 - 11:00)፦ **{count_opt1}** ድምጽ ({round(count_opt1/total_votes*100, 1)}%)\n"
        f"2️⃣ ከ 10:00 - 12:00 ይሁን፦ **{count_opt2}** ድምጽ ({round(count_opt2/total_votes*100, 1)}%)\n"
        f"━━━━━━━━━━━━━━━━━━━━━━"
    )
    await update.message.reply_text(stats_text, parse_mode="Markdown")
    # Send Excel file
    try:
        with open(SURVEY_EXCEL, "rb") as f:
            await update.message.reply_document(document=f, filename="survey_results.xlsx", caption="📄 የተሟላ የድምጽ ዝርዝር (Excel)")
    except Exception as e:
        logger.error(f"Could not send excel file: {e}")

# -------------------------------------------------------------
# 3. PRIVATE RESULT CHECKING CONVERSATION
# -------------------------------------------------------------
ASK_NAME, ASK_CLASS, ASK_PHONE = range(3)

async def start_private(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Entry point for private chat result checking or group intro."""
    # If in a group, send a guide
    if update.message.chat.type in ["group", "supergroup"]:
        await update.message.reply_text(
            "👋 ሰላም! ይህ የአጸደ ትጉሃን ቦት ነው።\n\n"
            "• የአስተያየት መጠይቅ በግሩፑ ለመጀመር 👉 /ask ይበሉ\n"
            "• የፈተና ውጤት ለማየት እባክዎ ቦቱን በግል (Private Chat) ያናግሩት 👉 @atsedeteguhandatacollect_bot"
        )
        return ConversationHandler.END

    # In Private Chat: Start result check
    await update.message.reply_text(
        "👋 እንኳን ወደ **አጸደ ትጉሃን ሰንበት ትምህርት ቤት** የውጤት መመልከቻ ቦት በደህና መጡ!\n\n"
        "እባክዎ የተማሪውን **ሙሉ ስም** ያስገቡ፦\n"
        "(ምሳሌ፦ አበበ ከበደ)",
        parse_mode="Markdown",
    )
    return ASK_NAME

async def get_student_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["student_name"] = update.message.text.strip()
    
    classes_keyboard = [["3ኛ ክፍል"], ["4ኛ ክፍል", "5ኛ ክፍል"]]
    await update.message.reply_text(
        f"✅ ስም፦ **{context.user_data['student_name']}**\n\nእባክዎ **ክፍል** ይምረጡ፦",
        reply_markup=ReplyKeyboardMarkup(classes_keyboard, one_time_keyboard=True, resize_keyboard=True),
        parse_mode="Markdown",
    )
    return ASK_CLASS

async def get_student_class(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["student_class"] = update.message.text.strip()
    
    # Phone Verification Keyboard
    contact_btn = [[KeyboardButton("📱 ስልክ ቁጥሬን አረጋግጥ (Share Phone)", request_contact=True)]]
    await update.message.reply_text(
        "🔒 **የደህንነት ማረጋገጫ (Verification)**\n\n"
        "ውጤትዎን ለመመልከት ከታች ያለውን **'📱 ስልክ ቁጥሬን አረጋግጥ'** የሚለውን ቁልፍ ይጫኑ ወይም ስልክ ቁጥርዎን በጽሁፍ ያስገቡ፦",
        reply_markup=ReplyKeyboardMarkup(contact_btn, one_time_keyboard=True, resize_keyboard=True),
        parse_mode="Markdown",
    )
    return ASK_PHONE

async def get_phone_and_show_result(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.contact:
        user_phone = update.message.contact.phone_number
    else:
        user_phone = update.message.text.strip()

    student_name = context.user_data.get("student_name", "").strip()
    student_class = context.user_data.get("student_class", "").strip()

    await update.message.reply_text("🔍 ውጤት በመፈለግ ላይ... እባክዎ ይጠብቁ...", reply_markup=ReplyKeyboardRemove())

    # Find Excel file matching class or default to grade 3.xlsx
    excel_files = [f for f in glob.glob(os.path.join(WORKING_DIR, "*.xlsx")) if not f.endswith("_backup.xlsx") and not "survey" in f and not "mail_merge" in f]
    
    if not excel_files:
        await update.message.reply_text("❌ የክፍሉ ውጤት መረጃ ፋይል አልተገኘም። እባክዎ አስተዳዳሪውን ያነጋግሩ።")
        return ConversationHandler.END

    src_excel = excel_files[0]
    wb = openpyxl.load_workbook(src_excel, data_only=True)
    ws = wb.active

    found_student = None

    for r in range(3, ws.max_row + 1):
        name_in_file = str(ws.cell(row=r, column=3).value or "").strip()
        if not name_in_file:
            continue

        # Check name match (fuzzy / partial)
        if student_name.lower() in name_in_file.lower() or name_in_file.lower() in student_name.lower():
            # Get scores
            scores = [ws.cell(row=r, column=c).value for c in range(6, 11)]
            valid_scores = [s for s in scores if isinstance(s, (int, float))]
            total = sum(valid_scores) if valid_scores else 0
            avg = round(total / len(valid_scores), 2) if valid_scores else 0
            
            found_student = {
                "name": name_in_file,
                "no": ws.cell(row=r, column=2).value or "",
                "scores": scores,
                "total": total,
                "avg": avg,
                "grade": ws.title if ws.title else student_class,
            }
            break

    if found_student:
        s1 = found_student["scores"][0] if len(found_student["scores"]) > 0 and found_student["scores"][0] is not None else "-"
        s2 = found_student["scores"][1] if len(found_student["scores"]) > 1 and found_student["scores"][1] is not None else "-"
        s3 = found_student["scores"][2] if len(found_student["scores"]) > 2 and found_student["scores"][2] is not None else "-"
        s4 = found_student["scores"][3] if len(found_student["scores"]) > 3 and found_student["scores"][3] is not None else "-"
        s5 = found_student["scores"][4] if len(found_student["scores"]) > 4 and found_student["scores"][4] is not None else "-"

        result_msg = (
            f"🎓 **አጸደ ትጉሃን ሰንበት ትምህርት ቤት**\n"
            f"📋 **የተማሪ ውጤት መግለጫ**\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"👤 **ስም፦** {found_student['name']}\n"
            f"🏫 **ክፍል፦** {found_student['grade']}\n"
            f"🔢 **መለያ ቁጥር፦** {found_student['no']}\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📖 ትምህርት 1፦ `{s1}`\n"
            f"📖 ትምህርት 2፦ `{s2}`\n"
            f"📖 ትምህርት 3፦ `{s3}`\n"
            f"📖 ትምህርት 4፦ `{s4}`\n"
            f"📖 ትምህርት 5፦ `{s5}`\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📊 **አጠቃላይ ድምር፦** `{found_student['total']}`\n"
            f"📈 **አማካይ ውጤት፦** `{found_student['avg']}%`\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"🎉 **መልካም ውጤት!**\n"
            f"እንደገና ውጤት ለማየት /start ይበሉ።"
        )
        await update.message.reply_text(result_msg, parse_mode="Markdown")
    else:
        await update.message.reply_text(
            f"❌ **ይቅርታ፦** '{student_name}' በሚል ስም ምንም የተመዘገበ ውጤት አልተገኘም።\n\n"
            f"እባክዎ ስሙን በትክክል ጽፈው እንደገና በ /start ይሞክሩ።"
        )

    return ConversationHandler.END

async def cancel_conversation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("የተሰረዘ ተግባር። እንደገና ለመጀመር /start ይጫኑ።", reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

# -------------------------------------------------------------
# MAIN APP
# -------------------------------------------------------------
def main():
    print("🚀 Bot starting...")
    init_survey_excel()
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    # Survey Handlers for Group
    # 1. English slash commands
    app.add_handler(CommandHandler(["ask", "survey"], cmd_ask))
    app.add_handler(CommandHandler(["stats", "result"], cmd_survey_stats))
    app.add_handler(CommandHandler(["reset", "clear"], cmd_reset_survey))
    
    # 2. Broad regex triggers (matches "መጠይቅ", "ጠይቅ", "ጥያቄ", "አስተያየት", "ውጤት", "አጽዳ" with/without slash and trailing spaces/mentions)
    amharic_survey_pattern = r'(?i)^\s*(/)?(ጠይቅ|መጠይቅ|ጥያቄ|አስተያየት|ask|survey)'
    amharic_stats_pattern = r'(?i)^\s*(/)?(ውጤት|ማጠቃለያ|ስታትስቲክስ|stats|result)'
    amharic_reset_pattern = r'(?i)^\s*(/)?(አጽዳ|አድስ|reset|clear)'
    
    app.add_handler(MessageHandler(filters.Regex(amharic_survey_pattern), cmd_ask))
    app.add_handler(MessageHandler(filters.Regex(amharic_stats_pattern), cmd_survey_stats))
    app.add_handler(MessageHandler(filters.Regex(amharic_reset_pattern), cmd_reset_survey))
    
    app.add_handler(CallbackQueryHandler(handle_survey_callback))

    # Private Chat Conversation for Student Results
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start_private)],
        states={
            ASK_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_student_name)],
            ASK_CLASS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_student_class)],
            ASK_PHONE: [
                MessageHandler(filters.CONTACT, get_phone_and_show_result),
                MessageHandler(filters.TEXT & ~filters.COMMAND, get_phone_and_show_result),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel_conversation)],
    )

    app.add_handler(conv_handler)
    
    port = int(os.environ.get("PORT", 0))
    render_url = os.environ.get("RENDER_EXTERNAL_URL", "https://atsede-bot.onrender.com")

    if port > 0:
        print(f"🚀 Running in Webhook Mode on port {port} at {render_url}...")
        app.run_webhook(
            listen="0.0.0.0",
            port=port,
            url_path="telegram",
            webhook_url=f"{render_url}/telegram",
            drop_pending_updates=True,
        )
    else:
        print("🚀 Running in Polling Mode...")
        app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
